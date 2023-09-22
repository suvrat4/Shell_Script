import requests
import json
from datetime import datetime, timedelta
from openpyxl import Workbook

# Set your Jira credentials and API endpoint
username = 'your_username'
password = 'your_password'
jira_url = 'https://your-jira-instance.atlassian.net/rest/api/2'

# Set the Jira project key and JQL query to filter issues for the past 15 days
project_key = 'YOUR_PROJECT_KEY'
query = f'project={project_key} AND created >= -15d'

# Create an Excel workbook
wb = Workbook()
ws = wb.active
ws.title = 'Jira Data'

# Define the headers for the Excel sheet
headers = ['Key', 'Summary', 'Assignee', 'Reporter', 'Created', 'Updated']

# Write headers to the Excel sheet
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Authenticate with Jira using basic authentication
auth = (username, password)

# Fetch data from Jira
try:
    response = requests.get(f'{jira_url}/search', params={'jql': query}, auth=auth)
    response.raise_for_status()

    issues = response.json().get('issues', [])
    
    for row_num, issue in enumerate(issues, 2):
        ws.cell(row=row_num, column=1, value=issue['key'])
        ws.cell(row=row_num, column=2, value=issue['fields']['summary'])
        ws.cell(row=row_num, column=3, value=issue['fields']['assignee']['displayName'])
        ws.cell(row=row_num, column=4, value=issue['fields']['reporter']['displayName'])
        
        created_date = datetime.strptime(issue['fields']['created'], '%Y-%m-%dT%H:%M:%S.%f%z')
        ws.cell(row=row_num, column=5, value=created_date.strftime('%Y-%m-%d %H:%M:%S'))
        
        updated_date = datetime.strptime(issue['fields']['updated'], '%Y-%m-%dT%H:%M:%S.%f%z')
        ws.cell(row=row_num, column=6, value=updated_date.strftime('%Y-%m-%d %H:%M:%S'))

    # Save the Excel file
    wb.save('jira_data.xlsx')
    print('Jira data saved to jira_data.xlsx')

except requests.exceptions.RequestException as e:
    print(f'Error: {e}')
