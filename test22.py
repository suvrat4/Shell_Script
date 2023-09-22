import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# Jira server URL and credentials
JIRA_SERVER = 'https://your-jira-instance.com'
USERNAME = 'your_username'
PASSWORD = 'your_password'

# Project key for the specific project you want to query
PROJECT_KEY = 'YOUR_PROJECT_KEY'

# Calculate the start and end dates for the last 15 days
end_date = datetime.now()
start_date = end_date - timedelta(days=14)  # 14 days ago

# Define the Jira API endpoint for searching issues
api_endpoint = f'{JIRA_SERVER}/rest/api/2/search'

# Initialize variables for pagination
start_at = 0
max_results = 1000  # Adjust as needed to fetch more results per page

# Initialize an empty list to store issue data
all_issue_data = []

while True:
    # Construct the JQL query for issues created in the last 15 days for the specified project with pagination
    jql_query = (
        f'project = {PROJECT_KEY} AND created >= "{start_date.strftime("%Y-%m-%d")}" '
        f'AND created <= "{end_date.strftime("%Y-%m-%d")}" ORDER BY Rank DESC'
    )

    # Set up headers for authentication
    headers = {
        'Content-Type': 'application/json',
    }

    # Create a session for making authenticated requests
    session = requests.Session()
    session.auth = (USERNAME, PASSWORD)

    # Define the query parameters including the fields you want (e.g., summary, assignee, status) and pagination
    params = {
        'jql': jql_query,
        'fields': 'summary,assignee,status',  # Add any additional fields you need
        'startAt': start_at,
        'maxResults': max_results,
    }

    # Make the API request to Jira
    response = session.get(api_endpoint, headers=headers, params=params)

    # Check if the request was successful
    if response.status_code == 200:
        data = response.json()
        issues = data.get('issues', [])

        # Create a DataFrame from the issues data for the current page
        issue_data = []
        for issue in issues:
            assignee_data = issue['fields']['assignee']
            assignee_name = assignee_data['displayName'] if assignee_data else 'Unassigned'
            issue_status = issue['fields']['status']['name'] if 'status' in issue['fields'] else 'Unknown'
            issue_data.append([issue["key"], issue["fields"]["summary"], assignee_name, issue_status])

        all_issue_data.extend(issue_data)

        # Check if there are more pages of results
        if len(issues) < max_results:
            break  # No more pages
        else:
            start_at += max_results  # Move to the next page
    else:
        print(f'Error: {response.status_code} - {response.text}')
        break

# Create a DataFrame with all issues
df = pd.DataFrame(all_issue_data, columns=["Issue Key", "Summary", "Assignee", "Status"])

# Save the DataFrame to an Excel file
excel_filename = "jira_issues.xlsx"
df.to_excel(excel_filename, sheet_name='All Issues', index=False)

# Calculate open and closed ticket counts
open_tickets = df[df['Status'] != 'Closed']
closed_tickets = df[df['Status'] == 'Closed']

# Create a DataFrame for open vs. closed ticket count
count_data = {
    'Open Tickets': [len(open_tickets)],
    'Closed Tickets': [len(closed_tickets)]
}

count_df = pd.DataFrame(count_data)

# Save the open vs. closed ticket count to a new tab in the Excel file
with pd.ExcelWriter(excel_filename, mode='a', engine='openpyxl') as writer:
    count_df.to_excel(writer, sheet_name='Open vs. Closed Count', index=False)

# Create a tab with a bar chart showing the count of issues for different statuses
wb = Workbook()
ws = wb.active
ws.title = 'Status Counts'

# Calculate counts for each status
status_counts = df['Status'].value_counts()

# Add data to the new tab
for idx, (status, count) in enumerate(status_counts.items(), start=1):
    ws.cell(row=idx, column=1, value=status)
    ws.cell(row=idx, column=2, value=count)

# Create a bar chart
chart = BarChart()
chart.title = 'Issue Status Counts'
chart.x_axis.title = 'Status'
chart.y_axis.title = 'Count'

data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=len(status_counts))
categories = Reference(ws, min_col=1, min_row=2, max_row=len(status_counts))
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Add the chart to the worksheet
ws.add_chart(chart, "D2")

# Save the Excel file
wb.save(excel_filename)

# Send the email with the Excel file as an attachment
email_subject = 'Jira Issues for the Last 15 Days'
email_from = 'your_email@example.com'
email_to = 'recipient@example.com'

message = MIMEMultipart()
message['From'] = email_from
message['To'] = email_to
message['Subject'] = email_subject

email_body = "Jira Issues for the Last 15 Days:\n\n"
message.attach(MIMEText(email_body, 'plain'))

with open(excel_filename, 'rb') as attachment:
    part = MIMEApplication(attachment.read(), Name=excel_filename)
    part['Content-Disposition'] = f'attachment; filename={excel_filename}'
    message.attach(part)

smtp_server = 'smtp.example.com'  # Replace with your SMTP server details
smtp_port = 587  # Replace with the appropriate port
smtp_username = 'your_email@example.com'
smtp_password = 'your_email_password'

try:
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
       
