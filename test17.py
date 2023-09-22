import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# Jira server URL and credentials
JIRA_SERVER = 'https://your-jira-instance.com'
USERNAME = 'your_username'
PASSWORD = 'your_password'

# Project key for the specific project you want to query
PROJECT_KEY = 'YOUR_PROJECT_KEY'

# Calculate the start and end dates for the last 15 days
end_date = datetime.now()
start_date = end_date - timedelta(days=14)  # 14 days ago

# Define the Jira API endpoint for searching issues
api_endpoint = f'{JIRA_SERVER}/rest/api/2/search'

# Initialize variables for pagination
start_at = 0
max_results = 1000  # Adjust as needed to fetch more results per page

# Initialize an empty list to store issue data
all_issue_data = []

while True:
    # Construct the JQL query for issues created in the last 15 days for the specified project with pagination
    jql_query = (
        f'project = {PROJECT_KEY} AND created >= -15d ORDER BY Rank DESC'
    )

    # Set up headers for authentication
    headers = {
        'Content-Type': 'application/json',
    }

    # Create a session for making authenticated requests
    session = requests.Session()
    session.auth = (USERNAME, PASSWORD)

    # Define the query parameters including the fields you want (e.g., summary, assignee, status, created, resolved) and pagination
    params = {
        'jql': jql_query,
        'fields': 'summary,assignee,status,created,resolutiondate',  # Add any additional fields you need
        'startAt': start_at,
        'maxResults': max_results,
    }

    # Make the API request to Jira
    response = session.get(api_endpoint, headers=headers, params=params)

    # Check if the request was successful
    if response.status_code == 200:
        data = response.json()
        issues = data.get('issues', [])

        # Create a DataFrame from the issues data for the current page
        issue_data = []
        for issue in issues:
            assignee_data = issue['fields']['assignee']
            assignee_name = assignee_data['displayName'] if assignee_data else 'Unassigned'
            issue_status = issue['fields']['status']['name'] if 'status' in issue['fields'] else 'Unknown'
            created_date = issue['fields']['created']
            resolved_date = issue['fields']['resolutiondate']
            
            if created_date and resolved_date:
                created_date = datetime.strptime(created_date, '%Y-%m-%dT%H:%M:%S.%f%z')
                resolved_date = datetime.strptime(resolved_date, '%Y-%m-%dT%H:%M:%S.%f%z')
                resolution_time = resolved_date - created_date
                mttr_hours = resolution_time.total_seconds() / 3600  # Convert to hours
            else:
                mttr_hours = None

            issue_data.append([issue["key"], issue["fields"]["summary"], assignee_name, issue_status, mttr_hours])

        all_issue_data.extend(issue_data)

        # Check if there are more pages of results
        if len(issues) < max_results:
            break  # No more pages
        else:
            start_at += max_results  # Move to the next page
    else:
        print(f'Error: {response.status_code} - {response.text}')
        break

# Create a DataFrame with all issues, including MTTR data
df = pd.DataFrame(all_issue_data, columns=["Issue Key", "Summary", "Assignee", "Status", "MTTR (hours)"])

# Save the DataFrame to an Excel file
excel_filename = "jira_issues.xlsx"
df.to_excel(excel_filename, sheet_name='Jira Issues', index=False)

# Calculate MTTR statistics
mttr_avg = df['MTTR (hours)'].mean()
mttr_median = df['MTTR (hours)'].median()

# Calculate assignee counts
assignee_counts = df['Assignee'].value_counts()

# Create a new tab for assignee counts
wb = Workbook()
ws1 = wb.active
ws1.title = 'Jira Issues'

# Add MTTR statistics to the new tab
ws1['A1'] = 'MTTR (Mean Time to Resolve) Statistics'
ws1['A2'] = 'Average MTTR (hours)'
ws1['B2'] = mttr_avg
ws1['A3'] = 'Median MTTR (hours)'
ws1['B3'] = mttr_median

# Add assignee counts to the new tab
ws1['D1'] = 'Assignee'
ws1['E1'] = 'Count'

for idx, (assignee, count) in enumerate(assignee_counts.items(), start=2):
    ws1.cell(row=idx, column=4, value=assignee)
    ws1.cell(row=idx, column=5, value=count)

# Create a bar chart for status counts
chart = BarChart()
chart.title = 'Issue Status Counts'
chart.x_axis.title = 'Status'
chart.y_axis.title = 'Count'

data = Reference(ws1, min_col=5, min_row=1, max_col=5, max_row=len(assignee_counts))
categories = Reference(ws1, min_col=4, min_row=2, max_row=len(assignee_counts))
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)

# Add the chart to the worksheet
ws1.add_chart(chart, "H2")

# Save the Excel file
wb.save(excel_filename)

# Send the email with the Excel file as an attachment
email_subject = 'Jira Issues for the Last 15 Days'
email_from = 'your_email@example.com'
email_to = 'recipient@example.com'

message = MIMEMultipart()
message['From'] = email_from
message['To'] = email_to
message['Subject'] = email_subject

email_body = "Jira Issues for the Last 15 Days:\n\n"
message.attach(MIMEText(email_body, 'plain'))

with open(excel_filename, 'rb') as attachment:
    part = MIMEApplication(attachment.read(), Name=excel_filename)
    part['Content-Disposition'] = f'attachment; filename={excel_filename}'
    message.attach(part)

smtp_server = 'smtp.example.com'  # Replace with your SMTP server details
smtp_port = 587  # Replace with the appropriate port
smtp_username = 'your_email@example.com'
smtp_password = 'your_email_password'

try:
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(smtp_username, smtp_password)
        server.sendmail(email_from, email_to, message.as_string())
    print("Email sent successfully.")
except Exception as e:
    print(f"Error sending
