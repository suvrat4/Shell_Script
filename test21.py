
import requests
import json
from datetime import datetime, timedelta
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Set your Jira credentials and API endpoint
username = 'your_username'
password = 'your_password'
jira_url = 'https://your-jira-instance.atlassian.net/rest/api/2'

# Set the Jira project key and JQL query to filter issues for the past 15 days
project_key = 'YOUR_PROJECT_KEY'
query = f'project={project_key} AND created >= -15d'

# Create an Excel workbook
wb = Workbook()
ws = wb.active
ws.title = 'Jira Data'

# Define the headers for the Excel sheet
headers = ['Key', 'Summary', 'Assignee', 'Reporter', 'Created', 'Updated']

# Write headers to the Excel sheet
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

# Authenticate with Jira using basic authentication
auth = (username, password)

start_at = 0
max_results = 100  # You can adjust this to your preferred batch size

while True:
    try:
        response = requests.get(
            f'{jira_url}/search',
            params={'jql': query, 'startAt': start_at, 'maxResults': max_results},
            auth=auth
        )
        response.raise_for_status()

        issues = response.json().get('issues', [])

        if not issues:
            break

        for row_num, issue in enumerate(issues, start_at + 2):
            ws.cell(row=row_num, column=1, value=issue['key'])
            ws.cell(row=row_num, column=2, value=issue['fields']['summary'])
            ws.cell(row=row_num, column=3, value=issue['fields']['assignee']['displayName'])
            ws.cell(row=row_num, column=4, value=issue['fields']['reporter']['displayName'])

            created_date = datetime.strptime(issue['fields']['created'], '%Y-%m-%dT%H:%M:%S.%f%z')
            ws.cell(row=row_num, column=5, value=created_date.strftime('%Y-%m-%d %H:%M:%S'))

            updated_date = datetime.strptime(issue['fields']['updated'], '%Y-%m-%dT%H:%M:%S.%f%z')
            ws.cell(row=row_num, column=6, value=updated_date.strftime('%Y-%m-%d %H:%M:%S'))

        # Move the starting point for the next batch
        start_at += max_results

    except requests.exceptions.RequestException as e:
        print(f'Error: {e}')
        break

# Save the Excel file
wb.save('jira_data.xlsx')
print('Jira data saved to jira_data.xlsx')

# Email configuration
sender_email = 'your_email@gmail.com'
receiver_email = 'recipient_email@example.com'
email_subject = 'Jira Data for the Past 15 Days'
email_body = 'Please find the attached Excel file with Jira data.'

# Create a multipart message
msg = MIMEMultipart()
msg['From'] = sender_email
msg['To'] = receiver_email
msg['Subject'] = email_subject

# Attach the Excel file to the email
attachment = open('jira_data.xlsx', 'rb')
part = MIMEBase('application', 'octet-stream')
part.set_payload(attachment.read())
encoders.encode_base64(part)
part.add_header('Content-Disposition', 'attachment; filename="jira_data.xlsx"')
msg.attach(part)

# Add the email body as plain text
msg.attach(MIMEText(email_body, 'plain'))

# Connect to the SMTP server (Gmail SMTP server used in this example)
smtp_server = 'smtp.gmail.com'
smtp_port = 587
smtp_username = 'your_email@gmail.com'
smtp_password = 'your_email_password'

try:
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(smtp_username, smtp_password)

    # Send the email
    server.sendmail(sender_email, receiver_email, msg.as_string())
    print('Email sent successfully')
except Exception as e:
    print(f'Error sending email: {str(e)}')
finally:
    # Close the SMTP server connection
    server.quit()
